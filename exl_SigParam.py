''' generate *.vparam for vTESTstudio parametor '''

import pandas as pd
import openpyxl

##### fun(): process data and generate data frame #####
# this function will remove Empty and Strikethrough format data
def process_data(workbook, sheet_name):
    worksheet = workbook[sheet_name]
    pData = []      # processed Data
    Sig_index = 6   # column index of 'Signal Name'
    Msg_index = 10  # column index of 'Message Name'
    Mab_index = 20  # column index of 'Mab'

    # for row_index in worksheet.iter_rows(values_only=True):
    for row_index in range(2, worksheet.max_row+1):
        Msg_value = worksheet.cell(row=row_index, column=Msg_index).value
        Mab_value = worksheet.cell(row=row_index, column=Mab_index).value
        Msg_strike = worksheet.cell(row=row_index, column=Msg_index).font.strike
        Sig_strike = worksheet.cell(row=row_index, column=Sig_index).font.strike

        # generate a processed Data
        if Msg_value and Mab_value != None and not Msg_strike and not Sig_strike:
            pData.append([worksheet.cell(row=row_index, column=col).value for col in range(1, worksheet.max_column + 1)])

    # get column name
    columns = [worksheet.cell(row=1, column=col).value for col in range(1, worksheet.max_column + 1)]

    # convert pData to DataFrame
    df = pd.DataFrame(pData, columns=columns)
    # convert 'Message ID' from Hex to Dec format
    df['Message ID'] = df['Message ID'].apply(lambda x: int(x, 16))
    return df

##### fun(): check for special cases in signal names #####
def chk_signalname(sigName):
    sigName = sigName.replace('(PS:自定义)','')
    sigName = sigName.replace(' \n(PS: 自定义)','')
    sigName = sigName.replace(' ','')
    if '\n' in sigName and not 'EMMC_BYTE_' in sigName:
        # if there are multiple signal names, take the last signal name
        sigName_split = sigName.split('\n')
        final_sigName = sigName_split[len(sigName_split)-1]
        print(sigName,'  --->  ', final_sigName)
    else:
        sigName = sigName.replace('\n','')
        final_sigName = sigName
    return final_sigName

# read Excel
excel_fileName = 'VIU_TR05_信號定義_V04_20240201.xlsx'
workbook = openpyxl.load_workbook(excel_fileName, data_only=True)

# pre-precess data and return dataframe
df_can01 = process_data(workbook, 'CAN01_Matrix')
df_can02 = process_data(workbook, 'CAN02_Matrix')
df_lin01 = process_data(workbook, 'LIN01_Matrix')


'''sourceData_CANin = pd.read_excel(excel_fileName, engine='openpyxl', sheet_name='CAN_MAP_IN')
sourceData_CANout = pd.read_excel(excel_fileName, engine='openpyxl', sheet_name='CAN_MAP_OUT')
sourceData_LINin = pd.read_excel(excel_fileName, engine='openpyxl', sheet_name='LIN_MAP_IN')
sourceData_LINout = pd.read_excel(excel_fileName, engine='openpyxl', sheet_name='LIN_MAP_OUT')'''

PreviousGroup = 'DefaultGroup'
output_text = 'Vector Parameter	1.0'

'''-----------------------------------------------------------'''
############## CAN01_Matrix content ##############
'''-----------------------------------------------------------'''
for index, row in df_can01.iterrows():
    MsgName = row['Message Name']
    #MsgID = row['Message ID']
    SigName = row['Signal Name']
    SigName = chk_signalname(SigName)
    Bus = 'CAN1'
    TxNode = row['Transmitter']
    #NameSpace = f'pC_{MsgName}_{MsgID}'
    NameSpace = f'pC_{MsgName}'

    if pd.notna(SigName) and pd.notna(MsgName):  # if 'SigName' and 'MsgName'  is not empty
        if PreviousGroup != MsgName:
            # create a new Parameter Group
            output_text += f'\nGroup	{MsgName}\n\n'
            output_text += 'ScalarSingleRecord\n\n'
            output_text += 'Name	Type	Info	Value\n'
            output_text += f'{NameSpace}::{SigName}_pC	Signal		{Bus}::{TxNode}::{MsgName}::{SigName}\n'
            PreviousGroup = MsgName
        else:
            output_text += f'{NameSpace}::{SigName}_pC	Signal		{Bus}::{TxNode}::{MsgName}::{SigName}\n'


'''-----------------------------------------------------------'''
############## CAN02_Matrix content ##############
'''-----------------------------------------------------------'''
for index, row in df_can02.iterrows():
    MsgName = row['Message Name']
    #MsgID = row['Message ID']
    SigName = row['Signal Name']
    SigName = chk_signalname(SigName)
    Bus = 'CAN2'
    TxNode = row['Transmitter']
    #NameSpace = f'pC_{MsgName}_{MsgID}'
    NameSpace = f'pC_{MsgName}'

    if pd.notna(SigName) and pd.notna(MsgName):  # if 'SigName' and 'MsgName'  is not empty
        if PreviousGroup != MsgName:
            # create a new Parameter Group
            output_text += f'\nGroup	{MsgName}\n\n'
            output_text += 'ScalarSingleRecord\n\n'
            output_text += 'Name	Type	Info	Value\n'
            output_text += f'{NameSpace}::{SigName}_pC	Signal		{Bus}::{TxNode}::{MsgName}::{SigName}\n'
            PreviousGroup = MsgName
        else:
            output_text += f'{NameSpace}::{SigName}_pC	Signal		{Bus}::{TxNode}::{MsgName}::{SigName}\n'


'''-----------------------------------------------------------'''
############## LIN01_Matrix content ##############
'''-----------------------------------------------------------'''
for index, row in df_lin01.iterrows():
    MsgName = row['Message Name']
    #MsgID = row['Message ID']
    SigName = row['Signal Name']
    SigName = chk_signalname(SigName)
    Bus = 'LIN1'
    TxNode = row['Transmitter']
    #NameSpace = f'pL_{MsgName}_{MsgID}'
    NameSpace = f'pL_{MsgName}'

    if pd.notna(SigName) and pd.notna(MsgName):  # if 'SigName' and 'MsgName'  is not empty
        if PreviousGroup != MsgName:
            # create a new Parameter Group
            output_text += f'\nGroup	{MsgName}\n\n'
            output_text += 'ScalarSingleRecord\n\n'
            output_text += 'Name	Type	Info	Value\n'
            output_text += f'{NameSpace}::{SigName}_pL	Signal		{Bus}::{TxNode}::{MsgName}::{SigName}\n'
            PreviousGroup = MsgName
        else:
            output_text += f'{NameSpace}::{SigName}_pL	Signal		{Bus}::{TxNode}::{MsgName}::{SigName}\n'

# create a *.vparam file
with open('vTS_SignalMapping.vparam', 'w', encoding='utf-8') as f:
    f.write(output_text)

print("vTS_SignalMapping.vparam is created")
